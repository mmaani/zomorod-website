#!/usr/bin/env python3
"""
Zomorod Supplier Intelligence Autofill (v2)

- Per-country / per-source harvesting (wave execution)
- Append mode (default) or replace mode
- Optional seed-file input (txt with URLs)
- Robust HTTP (retries, timeouts, SSL handshake protection)
- Email filtering + dedupe
"""
from __future__ import annotations

import argparse
import os
import re
import shutil
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from bs4 import FeatureNotFound
from openpyxl import load_workbook

UA = {
    "User-Agent": "Mozilla/5.0 (compatible; ZomorodRFQBot/1.0; +info@zomorodmedical.com)"
}

EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.I)

EMAIL_BLOCKLIST = {
    "customercare@medzell.net",
    "support@medzell.net",
    "info@medzell.net",
    "customercare@turkishhealthcare.org",
}

SOCIAL_DOMAINS = ("facebook.com", "linkedin.com", "instagram.com", "twitter.com", "t.me", "wa.me", "whatsapp.com")

COLUMNS = [
    "Supplier_ID","Company","Country","Supplier_Type",
    "Primary_Category","Secondary_Categories","Product_Focus (free text)",
    "Website","Email(s)","Phone/WhatsApp",
    "Export_Experience (explicit)",
    "Cert_ISO13485_Claim","Cert_CE_Claim","Cert_Other (FDA/UKCA/etc.)",
    "Evidence_URL (email or cert claim page)",
    "Expected_Price_Range_USD (sanity check)",
    "Risk_Score (0=low,100=high)","Risk_Level",
    "Status","Last_Checked","Notes"
]

WORKBOOK_TEMPLATE_NAME = "Zomorod_Supplier_Intelligence_TEMPLATE.xlsx"
WORKBOOK_LIVE_NAME = "Zomorod_Supplier_Intelligence_LIVE.xlsx"

COUNTRY_ALIASES = {
    "uae": "United Arab Emirates",
    "u.a.e": "United Arab Emirates",
    "united arab emirates": "United Arab Emirates",
    "ksa": "Saudi Arabia",
    "saudi": "Saudi Arabia",
    "saudi arabia": "Saudi Arabia",
    "jordan": "Jordan",
    "turkey": "Turkey",
    "turkiye": "Turkey",
    "china": "China",
    "india": "India",
}

TRUE_LIKE = {"1", "true", "yes", "y", "stated"}
FALSE_LIKE = {"0", "false", "no", "n", "none", "not stated"}

STATUS_INSERT_READY = "Insert_Ready"
STATUS_REVIEW_DUPLICATE = "Review_Duplicate"
STATUS_REJECTED_INVALID = "Rejected_Invalid"

@dataclass
class SourceCfg:
    name: str
    base_url: str
    strategy: str
    max_items: int
    max_pages: int
    delay_s: float

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("xlsx", nargs="?", default="", help="Workbook path. LIVE workbook is enforced by default.")
    p.add_argument("--country", default="", help="Only process Seed_URLs rows matching this Country (exact). Blank = all.")
    p.add_argument("--source", default="", help="Only process Seed_URLs rows matching this Source_Name (exact). Blank = all.")
    p.add_argument("--mode", default="append", choices=["append", "replace"], help="append (default) or replace")
    p.add_argument("--limit", type=int, default=0, help="Stop after collecting this many NEW suppliers (0 = no early stop)")
    p.add_argument("--seed-file", default="", help="Optional text file with URLs to crawl (one per line).")
    p.add_argument("--delay", type=float, default=0.0, help="Override delay between requests (seconds). 0 = use Source_Config delay.")
    p.add_argument("--timeout", type=int, default=30, help="HTTP timeout seconds.")
    p.add_argument("--retries", type=int, default=2, help="HTTP retries per URL.")
    return p.parse_args()

def resolve_live_workbook_path(requested_path: str) -> str:
    requested = Path(requested_path).expanduser().resolve() if requested_path else None

    if requested and requested.name == WORKBOOK_LIVE_NAME:
        live_path = requested
    elif requested and requested.name == WORKBOOK_TEMPLATE_NAME:
        live_path = requested.with_name(WORKBOOK_LIVE_NAME)
    elif requested:
        return str(requested)
    else:
        live_path = Path(__file__).resolve().parent.parent / "automation" / "input" / WORKBOOK_LIVE_NAME

    template_path = live_path.with_name(WORKBOOK_TEMPLATE_NAME)
    if not live_path.exists():
        if not template_path.exists():
            raise SystemExit(
                f"LIVE workbook not found: {live_path}; TEMPLATE not found: {template_path}"
            )
        shutil.copyfile(template_path, live_path)
        print(f"[init] Copied TEMPLATE workbook to LIVE workbook: {live_path}")

    return str(live_path)

def http_get(url: str, *, timeout: int, retries: int) -> Optional[str]:
    if timeout <= 0:
        timeout = 30
    for attempt in range(retries + 1):
        try:
            r = requests.get(url, headers=UA, timeout=timeout)
            r.raise_for_status()
            return r.text
        except requests.exceptions.SSLError:
            time.sleep(0.4 + attempt * 0.4)
            continue
        except Exception:
            time.sleep(0.3 + attempt * 0.5)
            continue
    return None

def normalize_obfuscations(text: str) -> str:
    t = (text or "")
    t = t.replace("[at]", "@").replace("(at)", "@").replace(" at ", "@")
    t = t.replace("[dot]", ".").replace("(dot)", ".").replace(" dot ", ".")
    return t

def extract_emails(html: str) -> List[str]:
    html2 = normalize_obfuscations(html or "")
    emails = sorted(set(EMAIL_RE.findall(html2)))
    emails = [e for e in emails if e.lower() not in EMAIL_BLOCKLIST]
    return emails

def normalize_text(v: object) -> str:
    return re.sub(r"\s+", " ", str(v or "")).strip()

def normalize_country(country: str) -> str:
    raw = normalize_text(country)
    if not raw:
        return ""
    key = raw.lower().replace(".", "")
    return COUNTRY_ALIASES.get(key, raw)

def normalize_email_list(value: str) -> str:
    raw = normalize_text(value)
    if not raw:
        return ""
    parts = [p.strip().lower() for p in re.split(r"[,;\s]+", raw) if p.strip()]
    valid = sorted({e for e in parts if EMAIL_RE.fullmatch(e) and e not in EMAIL_BLOCKLIST})
    return ", ".join(valid[:5])

def normalize_phone(value: str) -> str:
    raw = normalize_text(value)
    if not raw:
        return ""
    digits = re.sub(r"[^\d+]", "", raw)
    digits = re.sub(r"(?!^)\+", "", digits)
    return digits[:40]

def normalize_bool_like(value: str) -> str:
    raw = normalize_text(value).lower()
    if not raw:
        return ""
    if raw in TRUE_LIKE:
        return "Stated"
    if raw in FALSE_LIKE:
        return ""
    return normalize_text(value)

def normalize_url(value: str) -> str:
    raw = normalize_text(value)
    if not raw:
        return ""
    candidate = raw if raw.startswith(("http://", "https://")) else f"https://{raw}"
    try:
        u = urlparse(candidate)
        host = (u.netloc or "").lower().replace("www.", "")
        if not host:
            return ""
        path = u.path or "/"
        return f"https://{host}{path}"
    except Exception:
        return ""

def extract_title(soup: BeautifulSoup) -> str:
    for tag in ["h1", "h2"]:
        h = soup.find(tag)
        if h and h.get_text(strip=True):
            return h.get_text(" ", strip=True)
    if soup.title and soup.title.get_text(strip=True):
        return soup.title.get_text(" ", strip=True)
    return ""

def best_website_from_page(soup: BeautifulSoup, fallback_url: str) -> str:
    for a in soup.select("a[href]"):
        href = (a.get("href") or "").strip()
        if href.startswith("http"):
            lh = href.lower()
            if lh.startswith("mailto:") or lh.startswith("tel:"):
                continue
            if any(d in lh for d in SOCIAL_DOMAINS):
                continue
            return href
    p = urlparse(fallback_url)
    if p.scheme and p.netloc:
        return f"{p.scheme}://{p.netloc}/"
    return ""

def read_source_config(ws) -> Dict[str, SourceCfg]:
    cfg: Dict[str, SourceCfg] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        base_url = str(row[1] or "").strip()
        strategy = str(row[2] or "").strip()
        max_items = int(row[3] or 0)
        max_pages = int(row[4] or 0)
        delay_s = float(row[5] or 0.6)
        cfg[name] = SourceCfg(name, base_url, strategy, max_items, max_pages, delay_s)
    return cfg

def read_seed_urls(ws, *, country: str, source: str) -> List[Dict[str, str]]:
    seeds: List[Dict[str, str]] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        url = str(row[2] or "").strip()
        use_raw = row[5]
        if isinstance(use_raw, bool):
            use = "TRUE" if use_raw else "FALSE"
        else:
            use = str(use_raw or "").strip().upper()
        if use not in {"TRUE", "YES", "Y", "1"} or not url:
            continue
        src = str(row[0] or "").strip()
        cty = str(row[1] or "").strip()
        if country and cty != country:
            continue
        if source and src != source:
            continue
        seeds.append({
            "Source_Name": src,
            "Country": cty,
            "URL": url,
            "Category_Hint": str(row[3] or "").strip(),
            "Seed_Notes": str(row[4] or "").strip(),
        })
    return seeds

def read_keywords(ws) -> Dict[str, List[str]]:
    kw: Dict[str, List[str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        cat = str(row[0]).strip()
        kws = [k.strip().lower() for k in str(row[1] or "").split(",") if k.strip()]
        kw[cat] = kws
    return kw

def map_categories(text: str, keyword_map: Dict[str, List[str]]) -> Tuple[str, str]:
    t = (text or "").lower()
    hits = []
    for cat, kws in keyword_map.items():
        score = sum(1 for k in kws if k and k in t)
        if score > 0:
            hits.append((cat, score))
    if not hits:
        return ("", "")
    hits.sort(key=lambda x: x[1], reverse=True)
    primary = hits[0][0]
    secondary = ", ".join([c for c, _ in hits[1:]])
    return (primary, secondary)

def infer_supplier_type(text: str) -> str:
    t = (text or "").lower()
    exporter_words = ["exporter", "trading", "trader", "export", "distributor", "wholesale"]
    manufacturer_words = ["manufacturer", "manufacturing", "factory", "oem", "production"]
    exp = any(w in t for w in exporter_words)
    man = any(w in t for w in manufacturer_words)
    if man and not exp:
        return "Manufacturer"
    if exp and not man:
        return "Exporter"
    if man and exp:
        return "Manufacturer"
    return ""

def extract_cert_claims(text: str) -> Tuple[str, str, str]:
    t = (text or "").lower()
    iso = "Stated" if ("iso 13485" in t or "iso13485" in t) else ""
    ce = "Stated" if ("ce mark" in t or "ce certified" in t or " ce " in f" {t} ") else ""
    other = []
    for w in ["fda", "ukca", "gmp", "sgs", "tuv", "bv", "intertek"]:
        if w in t:
            other.append(w.upper())
    return iso, ce, ", ".join(sorted(set(other)))

def calc_risk(emails: List[str], iso: str, ce: str, evidence_url: str, supplier_type: str) -> int:
    score = 0
    em = " ".join(emails).lower()
    if (not emails) or any(dom in em for dom in ["gmail.com", "yahoo.com", "hotmail.com", "outlook.com"]):
        score += 20
    if not iso:
        score += 15
    if not ce:
        score += 10
    if not evidence_url:
        score += 15
    if supplier_type == "Exporter":
        score += 10
    return min(100, score)

def normalize_domain(url: str) -> str:
    try:
        p = urlparse(url)
        host = (p.netloc or "").lower().strip()
        host = host.replace("www.", "")
        return host
    except Exception:
        return ""

def get_target_rows(ws_out) -> int:
    n = 0
    for r in range(2, ws_out.max_row + 1):
        v = ws_out.cell(r, 1).value
        if isinstance(v, int) and v > 0:
            n = max(n, v)
    return n if n > 0 else 180

def load_existing_keys(ws_out, target: int) -> set:
    seen = set()
    for r in range(2, target + 2):
        company = str(ws_out.cell(r, 2).value or "").strip()
        emails = str(ws_out.cell(r, 9).value or "").strip()
        website = str(ws_out.cell(r, 8).value or "").strip()
        if not company and not emails:
            continue
        key = (company.lower(), normalize_domain(website), emails.lower())
        seen.add(key)
    return seen

def find_first_empty_row(ws_out, target: int) -> int:
    for r in range(2, target + 2):
        if not str(ws_out.cell(r, 2).value or "").strip() and not str(ws_out.cell(r, 9).value or "").strip():
            return r
    return target + 2

def clear_output(ws_out, target: int):
    for r in range(2, target + 2):
        for c in range(2, len(COLUMNS) + 1):
            if c == 18:
                continue
            ws_out.cell(r, c).value = None

def write_row(ws_out, r: int, row: Dict[str, object]):
    ws_out.cell(r, 2).value = row.get("Company", "")
    ws_out.cell(r, 3).value = row.get("Country", "")
    ws_out.cell(r, 4).value = row.get("Supplier_Type", "")
    ws_out.cell(r, 5).value = row.get("Primary_Category", "")
    ws_out.cell(r, 6).value = row.get("Secondary_Categories", "")
    ws_out.cell(r, 7).value = row.get("Product_Focus (free text)", "")
    ws_out.cell(r, 8).value = row.get("Website", "")
    ws_out.cell(r, 9).value = row.get("Email(s)", "")
    ws_out.cell(r, 10).value = row.get("Phone/WhatsApp", "")
    ws_out.cell(r, 12).value = row.get("Cert_ISO13485_Claim", "")
    ws_out.cell(r, 13).value = row.get("Cert_CE_Claim", "")
    ws_out.cell(r, 14).value = row.get("Cert_Other (FDA/UKCA/etc.)", "")
    ws_out.cell(r, 15).value = row.get("Evidence_URL (email or cert claim page)", "")
    ws_out.cell(r, 17).value = row.get("Risk_Score (0=low,100=high)", "")
    ws_out.cell(r, 18).value = row.get("Risk_Level", "")
    ws_out.cell(r, 19).value = row.get("Status", "New")
    ws_out.cell(r, 20).value = row.get("Last_Checked", "")
    ws_out.cell(r, 21).value = row.get("Notes", "")

def build_soup(html: str) -> BeautifulSoup:
    try:
        return BeautifulSoup(html, "lxml")
    except FeatureNotFound:
        return BeautifulSoup(html, "html.parser")

def crawl_urls(
    urls: List[Dict[str, str]],
    *,
    keywords: Dict[str, List[str]],
    default_delay_s: float,
    timeout: int,
    retries: int,
    limit: int,
    existing_keys: set,
    cfg: Dict[str, SourceCfg],
    delay_override: float
) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    for idx, s in enumerate(urls, 1):
        url = s["URL"]
        html = http_get(url, timeout=timeout, retries=retries)
        if not html:
            continue

        ems = extract_emails(html)
        if not ems:
            continue

        soup = build_soup(html)
        name = extract_title(soup) or url
        text = soup.get_text("\n", strip=True)

        primary, secondary = map_categories(text, keywords)
        stype = infer_supplier_type(text)
        iso, ce, other = extract_cert_claims(text)
        website = best_website_from_page(soup, url)

        company = name.strip()
        emails_str = ", ".join(ems[:5])
        key = (company.lower(), normalize_domain(website), emails_str.lower())
        if key in existing_keys:
            continue
        existing_keys.add(key)

        risk_score = calc_risk(ems, iso, ce, url, stype)
        row = {
            "Company": company,
            "Country": s.get("Country","").strip(),
            "Supplier_Type": stype,
            "Primary_Category": primary or s.get("Category_Hint",""),
            "Secondary_Categories": secondary,
            "Website": website,
            "Email(s)": emails_str,
            "Phone/WhatsApp": "",
            "Cert_ISO13485_Claim": iso,
            "Cert_CE_Claim": ce,
            "Cert_Other (FDA/UKCA/etc.)": other,
            "Evidence_URL (email or cert claim page)": url,
            "Risk_Score (0=low,100=high)": risk_score,
            "Risk_Level": "HIGH" if risk_score >= 60 else "MED" if risk_score >= 30 else "LOW",
            "Source_Name": s.get("Source_Name", ""),
            "Source_URL": url,
            "Status": "New",
            "Last_Checked": datetime.now(timezone.utc).date().isoformat(),
            "Notes": f"Seed={s.get('Source_Name','')}; {s.get('Seed_Notes','')}".strip(),
        }
        rows.append(row)

        if len(rows) % 25 == 0:
            print(f"✅ NEW suppliers: {len(rows)} (visited {idx}/{len(urls)} URLs)")

        if limit and len(rows) >= limit:
            break

        if delay_override > 0:
            time.sleep(delay_override)
        else:
            src_name = s.get("Source_Name", "")
            src_delay = cfg.get(src_name).delay_s if src_name in cfg else default_delay_s
            time.sleep(src_delay)
    return rows

def normalize_row(row: Dict[str, object]) -> Dict[str, object]:
    norm = dict(row)
    norm["Company"] = normalize_text(norm.get("Company", ""))
    norm["Country"] = normalize_country(str(norm.get("Country", "")))
    norm["Website"] = normalize_url(str(norm.get("Website", "")))
    norm["Email(s)"] = normalize_email_list(str(norm.get("Email(s)", "")))
    norm["Phone/WhatsApp"] = normalize_phone(str(norm.get("Phone/WhatsApp", "")))
    norm["Source_Name"] = normalize_text(norm.get("Source_Name", "")).title()
    norm["Source_URL"] = normalize_url(str(norm.get("Source_URL", "")))
    norm["Cert_ISO13485_Claim"] = normalize_bool_like(str(norm.get("Cert_ISO13485_Claim", "")))
    norm["Cert_CE_Claim"] = normalize_bool_like(str(norm.get("Cert_CE_Claim", "")))
    norm["Primary_Category"] = normalize_text(norm.get("Primary_Category", ""))
    norm["Secondary_Categories"] = normalize_text(norm.get("Secondary_Categories", ""))
    norm["Notes"] = normalize_text(norm.get("Notes", ""))
    return norm

def validate_row(row: Dict[str, object]) -> List[str]:
    reasons: List[str] = []
    if not row.get("Company"):
        reasons.append("missing_company")
    if not row.get("Country"):
        reasons.append("missing_country")
    if not row.get("Primary_Category"):
        reasons.append("missing_primary_category")
    if not row.get("Website") and not row.get("Email(s)"):
        reasons.append("missing_contact_point")
    if row.get("Email(s)") and not normalize_email_list(str(row.get("Email(s)"))):
        reasons.append("invalid_email_format")
    return reasons

def duplicate_flags(row: Dict[str, object], seen: Dict[str, set]) -> List[str]:
    company = str(row.get("Company", "")).lower()
    domain = normalize_domain(str(row.get("Website", "")))
    emails = str(row.get("Email(s)", "")).lower()
    phone = str(row.get("Phone/WhatsApp", ""))
    source_url = str(row.get("Source_URL", "")).lower()

    keys = {
        "name_domain": f"{company}|{domain}" if company and domain else "",
        "name_email": f"{company}|{emails}" if company and emails else "",
        "domain_email": f"{domain}|{emails}" if domain and emails else "",
        "phone": phone,
        "source_url": source_url,
    }

    hits: List[str] = []
    for k, v in keys.items():
        if v and v in seen[k]:
            hits.append(k)
    for k, v in keys.items():
        if v:
            seen[k].add(v)
    return hits

def enforce_supplier_pipeline(rows: List[Dict[str, object]]) -> Tuple[List[Dict[str, object]], Dict[str, int]]:
    seen = {"name_domain": set(), "name_email": set(), "domain_email": set(), "phone": set(), "source_url": set()}
    out: List[Dict[str, object]] = []
    stats = {STATUS_INSERT_READY: 0, STATUS_REVIEW_DUPLICATE: 0, STATUS_REJECTED_INVALID: 0}

    for row in rows:
        norm = normalize_row(row)
        reasons = validate_row(norm)
        if reasons:
            norm["Status"] = STATUS_REJECTED_INVALID
            norm["Notes"] = f"{norm.get('Notes','')} | Validation={','.join(reasons)}".strip(" |")
            stats[STATUS_REJECTED_INVALID] += 1
            out.append(norm)
            continue

        hits = duplicate_flags(norm, seen)
        if len(hits) >= 2 or any(h in {"domain_email", "name_email", "name_domain"} for h in hits):
            norm["Status"] = STATUS_REVIEW_DUPLICATE
            norm["Notes"] = f"{norm.get('Notes','')} | DuplicateSignals={','.join(hits)}".strip(" |")
            stats[STATUS_REVIEW_DUPLICATE] += 1
        else:
            norm["Status"] = STATUS_INSERT_READY
            stats[STATUS_INSERT_READY] += 1

        out.append(norm)

    return out, stats

def load_seed_file_urls(path: str) -> List[str]:
    if not path:
        return []
    if not os.path.exists(path):
        raise SystemExit(f"Seed file not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        out = []
        for ln in f:
            u = ln.strip()
            if u.startswith("http"):
                out.append(u)
        return out

def require_sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    raise SystemExit(f"Missing required sheet: {name}")

def main():
    args = parse_args()
    workbook_path = resolve_live_workbook_path(args.xlsx)

    if not os.path.exists(workbook_path):
        raise SystemExit(f"Workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path)
    ws_cfg = require_sheet(wb, "Source_Config")
    ws_seed = require_sheet(wb, "Seed_URLs")
    ws_kw = require_sheet(wb, "Category_Keywords")
    ws_out = require_sheet(wb, "Supplier_Intelligence")
    ws_log = wb["Run_Log"] if "Run_Log" in wb.sheetnames else wb.create_sheet("Run_Log")
    if ws_log.max_row == 1 and not any(ws_log.iter_rows(min_row=1, max_row=1, values_only=True)):
        ws_log.append(["timestamp", "written", "mode", "notes"])

    cfg = read_source_config(ws_cfg)
    keywords = read_keywords(ws_kw)
    seeds = read_seed_urls(ws_seed, country=args.country, source=args.source)

    if args.seed_file:
        seed_urls = load_seed_file_urls(args.seed_file)
        for u in seed_urls:
            seeds.append({
                "Source_Name": args.source or "SeedFile",
                "Country": args.country or "",
                "URL": u,
                "Category_Hint": "",
                "Seed_Notes": f"Seed-file: {args.seed_file}",
            })

    if not seeds:
        raise SystemExit("No seeds found. Set Use=TRUE in Seed_URLs (and/or pass --seed-file) and try again.")

    target = get_target_rows(ws_out)

    if args.mode == "replace":
        clear_output(ws_out, target)

    existing_keys = load_existing_keys(ws_out, target)

    default_delay = 0.6
    if args.source and args.source in cfg:
        default_delay = cfg[args.source].delay_s
    elif not args.source:
        default_delay = 0.8

    new_rows = crawl_urls(
        seeds,
        keywords=keywords,
        default_delay_s=default_delay,
        timeout=args.timeout,
        retries=args.retries,
        limit=args.limit,
        existing_keys=existing_keys,
        cfg=cfg,
        delay_override=args.delay,
    )

    pipeline_rows, pipeline_stats = enforce_supplier_pipeline(new_rows)

    written = 0
    if args.mode == "replace":
        r = 2
        for row in pipeline_rows:
            if r > target + 1:
                break
            write_row(ws_out, r, row)
            written += 1
            r += 1
    else:
        r = find_first_empty_row(ws_out, target)
        for row in pipeline_rows:
            if r > target + 1:
                break
            write_row(ws_out, r, row)
            written += 1
            r += 1

    effective_delay = args.delay if args.delay > 0 else default_delay

    ws_log.append([
        datetime.now(timezone.utc).isoformat(),
        written,
        f"{args.source or 'MULTI'}|{args.country or 'ALL'}|{args.mode}",
        (
            f"Visited {len(seeds)} seeds; wrote {written} rows. "
            f"insert_ready={pipeline_stats[STATUS_INSERT_READY]} "
            f"review_duplicate={pipeline_stats[STATUS_REVIEW_DUPLICATE]} "
            f"rejected_invalid={pipeline_stats[STATUS_REJECTED_INVALID]}. "
            f"limit={args.limit or 'none'} delay={effective_delay}s"
        )
    ])

    wb.save(workbook_path)
    print(f"✅ Saved: {workbook_path}")
    print(
        "✅ Pipeline summary: "
        f"insert_ready={pipeline_stats[STATUS_INSERT_READY]}, "
        f"review_duplicate={pipeline_stats[STATUS_REVIEW_DUPLICATE]}, "
        f"rejected_invalid={pipeline_stats[STATUS_REJECTED_INVALID]}"
    )
    print(f"✅ Wrote suppliers: {written}")

if __name__ == "__main__":
    main()
