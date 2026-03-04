#!/usr/bin/env python3
import re
import time
import sys
from dataclasses import dataclass
from typing import Dict, List, Tuple
from urllib.parse import urlparse
from datetime import datetime  # ✅ FIX

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

UA = {
    "User-Agent": "Mozilla/5.0 (compatible; ZomorodRFQBot/1.0; +info@zomorodmedical.com)"
}

EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.I)

COLUMNS = [
    "Supplier_ID","Company","Country","Supplier_Type",
    "Primary_Category","Secondary_Categories","Product_Focus (free text)",
    "Website","Email(s)","Phone/WhatsApp",
    "Export_Experience (explicit)",
    "Cert_ISO13485_Claim","Cert_CE_Claim","Cert_Other (FDA/UKCA/etc.)",
    "Evidence_URL (email or cert claim page)",
    "Expected_Price_Range_USD (sanity check)",
    "Risk_Score (0=low,100=high)","Risk_Level",
    "Status","Last_Checked","Notes"
]

@dataclass
class SourceCfg:
    name: str
    base_url: str
    strategy: str
    max_items: int
    max_pages: int
    delay_s: float

def http_get(url: str, timeout=30) -> str:
    r = requests.get(url, headers=UA, timeout=timeout)
    r.raise_for_status()
    return r.text

def normalize_obfuscations(text: str) -> str:
    t = text.replace("[at]", "@").replace("(at)", "@").replace(" at ", "@")
    t = t.replace("[dot]", ".").replace("(dot)", ".").replace(" dot ", ".")
    return t

def extract_emails(html: str) -> List[str]:
    html2 = normalize_obfuscations(html or "")
    return sorted(set(EMAIL_RE.findall(html2)))

def extract_title(soup: BeautifulSoup) -> str:
    for tag in ["h1", "h2"]:
        h = soup.find(tag)
        if h and h.get_text(strip=True):
            return h.get_text(" ", strip=True)
    if soup.title and soup.title.get_text(strip=True):
        return soup.title.get_text(" ", strip=True)
    return ""

def best_website_from_page(soup: BeautifulSoup, fallback_url: str) -> str:
    # Try to find a company website link; else fallback to page origin
    for a in soup.select("a[href]"):
        href = (a.get("href") or "").strip()
        if href.startswith("http"):
            if any(x in href.lower() for x in ["facebook", "linkedin", "instagram", "twitter", "t.me", "whatsapp"]):
                continue
            return href
    p = urlparse(fallback_url)
    if p.scheme and p.netloc:
        return f"{p.scheme}://{p.netloc}/"
    return ""

def read_source_config(ws) -> Dict[str, SourceCfg]:
    cfg = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        base_url = str(row[1] or "").strip()
        strategy = str(row[2] or "").strip()
        max_items = int(row[3] or 0)
        max_pages = int(row[4] or 0)
        delay_s = float(row[5] or 0.4)
        cfg[name] = SourceCfg(name, base_url, strategy, max_items, max_pages, delay_s)
    return cfg

def read_seed_urls(ws) -> List[Dict[str, str]]:
    seeds = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        url = str(row[2] or "").strip()
        use = str(row[5] or "").strip().upper()
        if use != "TRUE" or not url:
            continue
        seeds.append({
            "Source_Name": str(row[0] or "").strip(),
            "Country": str(row[1] or "").strip(),
            "URL": url,
            "Category_Hint": str(row[3] or "").strip(),
            "Seed_Notes": str(row[4] or "").strip(),
        })
    return seeds

def read_keywords(ws) -> Dict[str, List[str]]:
    kw = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        cat = str(row[0]).strip()
        kws = [k.strip().lower() for k in str(row[1] or "").split(",") if k.strip()]
        kw[cat] = kws
    return kw

def map_categories(text: str, keyword_map: Dict[str, List[str]]) -> Tuple[str, str]:
    t = (text or "").lower()
    hits = []
    for cat, kws in keyword_map.items():
        score = sum(1 for k in kws if k in t)
        if score > 0:
            hits.append((cat, score))
    if not hits:
        return ("", "")
    hits.sort(key=lambda x: x[1], reverse=True)
    primary = hits[0][0]
    secondary = ", ".join([c for c, _ in hits[1:]])
    return (primary, secondary)

def infer_supplier_type(text: str) -> str:
    t = (text or "").lower()
    exporter_words = ["exporter", "trading", "trader", "export", "distributor", "wholesale"]
    manufacturer_words = ["manufacturer", "manufacturing", "factory", "oem", "production"]
    exp = any(w in t for w in exporter_words)
    man = any(w in t for w in manufacturer_words)
    if man and not exp:
        return "Manufacturer"
    if exp and not man:
        return "Exporter"
    if man and exp:
        return "Manufacturer"
    return ""

def extract_cert_claims(text: str) -> Tuple[str, str, str]:
    t = (text or "").lower()
    iso = "Stated" if ("iso 13485" in t or "iso13485" in t) else ""
    ce = "Stated" if ("ce mark" in t or "ce certified" in t or " ce " in f" {t} ") else ""
    other = []
    for w in ["fda", "ukca", "gmp", "sgs", "tuv", "bv", "intertek"]:
        if w in t:
            other.append(w.upper())
    return iso, ce, ", ".join(sorted(set(other)))

def calc_risk(emails: List[str], iso: str, ce: str, evidence_url: str, supplier_type: str) -> int:
    score = 0
    em = " ".join(emails).lower()
    if (not emails) or ("gmail.com" in em) or ("yahoo" in em):
        score += 20
    if not iso:
        score += 15
    if not ce:
        score += 10
    if not evidence_url:
        score += 15
    if supplier_type == "Exporter":
        score += 10
    return min(100, score)

def clear_output(ws):
    # Keep Supplier_ID + Risk_Level formulas, clear others for rows 2..181
    for r in range(2, 182):
        for c in range(2, len(COLUMNS) + 1):
            # Column 18 (Risk_Level) has formula in your design; keep it
            if c == 18:
                continue
            ws.cell(r, c).value = None

def write_rows(ws, rows: List[Dict[str, str]]):
    max_rows = min(180, len(rows))
    for i in range(max_rows):
        r = i + 2
        row = rows[i]
        ws.cell(r, 2).value = row.get("Company", "")
        ws.cell(r, 3).value = row.get("Country", "")
        ws.cell(r, 4).value = row.get("Supplier_Type", "")
        ws.cell(r, 5).value = row.get("Primary_Category", "")
        ws.cell(r, 6).value = row.get("Secondary_Categories", "")
        ws.cell(r, 7).value = row.get("Product_Focus (free text)", "")
        ws.cell(r, 8).value = row.get("Website", "")
        ws.cell(r, 9).value = row.get("Email(s)", "")
        ws.cell(r, 10).value = row.get("Phone/WhatsApp", "")
        ws.cell(r, 11).value = row.get("Export_Experience (explicit)", "")
        ws.cell(r, 12).value = row.get("Cert_ISO13485_Claim", "")
        ws.cell(r, 13).value = row.get("Cert_CE_Claim", "")
        ws.cell(r, 14).value = row.get("Cert_Other (FDA/UKCA/etc.)", "")
        ws.cell(r, 15).value = row.get("Evidence_URL (email or cert claim page)", "")
        ws.cell(r, 16).value = row.get("Expected_Price_Range_USD (sanity check)", "")
        ws.cell(r, 17).value = row.get("Risk_Score (0=low,100=high)", "")
        ws.cell(r, 19).value = row.get("Status", "New")
        ws.cell(r, 20).value = row.get("Last_Checked", "")
        ws.cell(r, 21).value = row.get("Notes", "")

def dedupe(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    seen = set()
    out = []
    for r in rows:
        key = (r.get("Company","").strip().lower(), r.get("Email(s)","").strip().lower())
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out

def crawl_seed_urls(seeds, keywords, delay_s):
    rows = []
    for s in seeds:
        url = s["URL"]
        try:
            html = http_get(url)
            ems = extract_emails(html)
            if not ems:
                continue
            soup = BeautifulSoup(html, "lxml")
            name = extract_title(soup) or url
            text = soup.get_text("\n", strip=True)
            primary, secondary = map_categories(text, keywords)
            stype = infer_supplier_type(text)
            iso, ce, other = extract_cert_claims(text)
            website = best_website_from_page(soup, url)

            country = s.get("Country","").strip()

            row = {
                "Company": name,
                "Country": country,
                "Supplier_Type": stype,
                "Primary_Category": primary or s.get("Category_Hint",""),
                "Secondary_Categories": secondary,
                "Website": website,
                "Email(s)": ", ".join(ems[:5]),
                "Cert_ISO13485_Claim": iso,
                "Cert_CE_Claim": ce,
                "Cert_Other (FDA/UKCA/etc.)": other,
                "Evidence_URL (email or cert claim page)": url,
                "Risk_Score (0=low,100=high)": calc_risk(ems, iso, ce, url, stype),
                "Status": "New",
                "Last_Checked": datetime.utcnow().date().isoformat(),
                "Notes": f"Seed={s.get('Source_Name','')}; {s.get('Seed_Notes','')}".strip(),
            }
            rows.append(row)
            time.sleep(delay_s)
        except Exception:
            continue
    return rows

def main(xlsx_path: str):
    wb = load_workbook(xlsx_path)
    ws_cfg = wb["Source_Config"]
    ws_seed = wb["Seed_URLs"]
    ws_kw = wb["Category_Keywords"]
    ws_out = wb["Supplier_Intelligence"]
    ws_log = wb["Run_Log"]

    cfg = read_source_config(ws_cfg)
    keywords = read_keywords(ws_kw)
    seeds = read_seed_urls(ws_seed)

    clear_output(ws_out)

    # Only seed-driven is guaranteed; your listing sources can be added later
    delay = 0.4
    if "Medzell" in cfg:
        delay = cfg["Medzell"].delay_s

    rows = crawl_seed_urls(seeds, keywords, delay)
    rows = dedupe([r for r in rows if r.get("Email(s)") and r.get("Evidence_URL (email or cert claim page)")])
    rows = rows[:180]

    write_rows(ws_out, rows)

    ws_log.append([datetime.utcnow().isoformat()+"Z", len(rows), "Seed_URLs", f"Visited {len(seeds)} seeds; wrote {len(rows)} email-verified rows."])

    wb.save(xlsx_path)
    print(f"✅ Updated workbook saved: {xlsx_path}")
    print(f"✅ Rows filled: {len(rows)}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python scripts/zomorod_autofill_supplier_intelligence.py input/Zomorod_Supplier_Intelligence_180_AUTOFILL.xlsx")
        sys.exit(1)
    main(sys.argv[1])