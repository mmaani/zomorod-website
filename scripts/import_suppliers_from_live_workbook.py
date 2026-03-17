#!/usr/bin/env python3
"""
Import Insert_Ready supplier rows from LIVE workbook into /api/suppliers.

Safety defaults:
- Dry-run by default (no writes)
- Only imports rows with Status == Insert_Ready
- Blocks rows marked Review_Duplicate or Rejected_Invalid
- Runs duplicate safeguards against existing suppliers before POST
"""
from __future__ import annotations

import argparse
import json
import os
import re
import shutil
from pathlib import Path
from typing import Dict, List, Tuple

import requests
from openpyxl import load_workbook

WORKBOOK_TEMPLATE_NAME = "Zomorod_Supplier_Intelligence_TEMPLATE.xlsx"
WORKBOOK_LIVE_NAME = "Zomorod_Supplier_Intelligence_LIVE.xlsx"

STATUS_INSERT_READY = "insert_ready"
STATUS_REVIEW_DUPLICATE = "review_duplicate"
STATUS_REJECTED_INVALID = "rejected_invalid"

WORKFLOW_STATUS_DEFAULT = "UNDER_REVIEW"
VALID_RISK = {"LOW", "MED", "HIGH"}


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--base-url", default=os.getenv("BASE_URL", "http://localhost:3000"))
    p.add_argument("--auth-token", default=os.getenv("AUTH_TOKEN", ""))
    p.add_argument("--xlsx", default="", help="Workbook path. Defaults to enforced LIVE workbook path.")
    p.add_argument("--apply", action="store_true", help="Actually call POST /api/suppliers. Default is dry-run.")
    p.add_argument("--limit", type=int, default=0, help="Max eligible rows to process. 0 means no limit.")
    p.add_argument("--verbose", action="store_true")
    return p.parse_args()


def resolve_live_workbook_path(requested_path: str) -> str:
    requested = Path(requested_path).expanduser().resolve() if requested_path else None

    if requested and requested.name == WORKBOOK_LIVE_NAME:
        live_path = requested
    elif requested and requested.name == WORKBOOK_TEMPLATE_NAME:
        live_path = requested.with_name(WORKBOOK_LIVE_NAME)
    elif requested:
        return str(requested)
    else:
        live_path = Path(__file__).resolve().parent.parent / "automation" / "input" / WORKBOOK_LIVE_NAME

    template_path = live_path.with_name(WORKBOOK_TEMPLATE_NAME)
    if not live_path.exists():
        if not template_path.exists():
            raise SystemExit(
                f"LIVE workbook not found: {live_path}; TEMPLATE not found: {template_path}"
            )
        shutil.copyfile(template_path, live_path)
        print(f"[init] Copied TEMPLATE workbook to LIVE workbook: {live_path}")

    return str(live_path)


def normalize_text(v: object) -> str:
    return re.sub(r"\s+", " ", str(v or "")).strip()


def normalize_status(v: object) -> str:
    return normalize_text(v).lower()


def normalize_domain(url: str) -> str:
    s = normalize_text(url).lower()
    s = re.sub(r"^https?://", "", s)
    s = s.split("/")[0]
    return s.replace("www.", "")


def first_email(raw: str) -> str:
    for part in re.split(r"[,;\s]+", normalize_text(raw).lower()):
        if re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", part):
            return part
    return ""


def parse_source_name(notes: str) -> str:
    m = re.search(r"(?:^|\s)Seed\s*=\s*([^;|]+)", notes or "", flags=re.I)
    return normalize_text(m.group(1)) if m else "Automation Workbook"


def map_risk(v: object) -> str:
    up = normalize_text(v).upper()
    if up in VALID_RISK:
        return up
    return "MED"


def load_sheet_rows(path: str) -> List[Dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    if "Supplier_Intelligence" not in wb.sheetnames:
        raise SystemExit("Missing required sheet: Supplier_Intelligence")
    ws = wb["Supplier_Intelligence"]

    rows: List[Dict[str, str]] = []
    for r in range(2, ws.max_row + 1):
        row = {
            "Company": normalize_text(ws.cell(r, 2).value),
            "Country": normalize_text(ws.cell(r, 3).value),
            "Supplier_Type": normalize_text(ws.cell(r, 4).value),
            "Primary_Category": normalize_text(ws.cell(r, 5).value),
            "Secondary_Categories": normalize_text(ws.cell(r, 6).value),
            "Website": normalize_text(ws.cell(r, 8).value),
            "Email(s)": normalize_text(ws.cell(r, 9).value),
            "Phone/WhatsApp": normalize_text(ws.cell(r, 10).value),
            "Cert_ISO13485_Claim": normalize_text(ws.cell(r, 12).value),
            "Cert_CE_Claim": normalize_text(ws.cell(r, 13).value),
            "Cert_Other": normalize_text(ws.cell(r, 14).value),
            "Evidence_URL": normalize_text(ws.cell(r, 15).value),
            "Expected_Price_Range_USD": normalize_text(ws.cell(r, 16).value),
            "Risk_Level": normalize_text(ws.cell(r, 18).value),
            "Status": normalize_text(ws.cell(r, 19).value),
            "Last_Checked": normalize_text(ws.cell(r, 20).value),
            "Notes": normalize_text(ws.cell(r, 21).value),
            "_row": str(r),
        }
        if any(row.values()):
            rows.append(row)

    return rows


def req(base_url: str, token: str, method: str, path: str, payload=None):
    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    url = base_url.rstrip("/") + path
    try:
        res = requests.request(method, url, headers=headers, data=json.dumps(payload) if payload is not None else None, timeout=30)
    except requests.RequestException as exc:
        raise SystemExit(f"API request failed: {method} {url} :: {exc}")
    data = {}
    try:
        data = res.json()
    except Exception:
        pass
    return res.status_code, data


def category_maps(categories: List[Dict[str, object]]) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for c in categories:
        name = normalize_text(c.get("name", "")).lower()
        cid = int(c.get("id", 0) or 0)
        if name and cid > 0:
            out[name] = cid
    return out


def parse_secondary_ids(raw: str, category_by_name: Dict[str, int], primary_id: int) -> List[int]:
    ids: List[int] = []
    seen = set()
    for part in [normalize_text(x).lower() for x in raw.split(",") if normalize_text(x)]:
        cid = category_by_name.get(part)
        if cid and cid != primary_id and cid not in seen:
            seen.add(cid)
            ids.append(cid)
    return ids


def duplicate_signal(payload: Dict[str, object], existing: List[Dict[str, object]]) -> str:
    p_name = normalize_text(payload.get("legalName", "")).lower()
    p_country = normalize_text(payload.get("supplierCountry", "")).lower()
    p_email = normalize_text(payload.get("email", "")).lower()
    p_domain = normalize_domain(str(payload.get("website", "")))

    for s in existing:
        s_name = normalize_text(s.get("legalName") or s.get("businessName") or s.get("name")).lower()
        s_country = normalize_text(s.get("supplierCountry")).lower()
        s_email = normalize_text(s.get("email")).lower()
        s_domain = normalize_domain(str(s.get("website", "")))

        if p_name and p_country and p_name == s_name and p_country == s_country:
            return "name_country"
        if p_email and s_email and p_email == s_email:
            return "email"
        if p_domain and s_domain and p_domain == s_domain and p_name and p_name == s_name:
            return "name_domain"
    return ""


def build_payload(row: Dict[str, str], category_by_name: Dict[str, int]) -> Tuple[Dict[str, object], List[str]]:
    errors: List[str] = []

    legal_name = row["Company"]
    supplier_country = row["Country"]
    primary_cat = row["Primary_Category"].lower()
    primary_id = category_by_name.get(primary_cat, 0)

    if not legal_name:
        errors.append("missing_legal_name")
    if not supplier_country:
        errors.append("missing_country")
    if not primary_id:
        errors.append("missing_or_unmapped_primary_category")

    payload: Dict[str, object] = {
        "legalName": legal_name,
        "businessName": legal_name,
        "name": legal_name,
        "contactName": "",
        "phone": row["Phone/WhatsApp"],
        "phoneWhatsapp": row["Phone/WhatsApp"],
        "email": first_email(row["Email(s)"]),
        "website": row["Website"],
        "supplierCountry": supplier_country,
        "supplierCity": "",
        "supplierType": row["Supplier_Type"],
        "workflowStatus": WORKFLOW_STATUS_DEFAULT,
        "riskLevel": map_risk(row["Risk_Level"]),
        "certificationsIso13485": row["Cert_ISO13485_Claim"],
        "certificationsCe": row["Cert_CE_Claim"],
        "certificationsOther": row["Cert_Other"],
        "evidenceUrl": row["Evidence_URL"],
        "expectedPriceRangeUsd": row["Expected_Price_Range_USD"],
        "sourceName": parse_source_name(row["Notes"]),
        "sourceUrl": row["Evidence_URL"],
        "notes": f"Imported from Supplier_Intelligence row {row['_row']}. {row['Notes']}",
        "primaryCategoryId": primary_id,
        "secondaryCategoryIds": parse_secondary_ids(row["Secondary_Categories"], category_by_name, primary_id),
    }

    return payload, errors


def main() -> None:
    args = parse_args()
    workbook_path = resolve_live_workbook_path(args.xlsx)
    rows = load_sheet_rows(workbook_path)

    status, data = req(args.base_url, args.auth_token, "GET", "/api/suppliers?limit=2000")
    if status != 200 or not data.get("ok"):
        raise SystemExit(f"Failed to load categories/suppliers from API: status={status} body={data}")

    category_by_name = category_maps(data.get("categories") or [])
    existing_suppliers = data.get("suppliers") or []

    summary = {
        "total_rows": len(rows),
        "eligible_insert_ready": 0,
        "blocked_review_duplicate": 0,
        "blocked_rejected_invalid": 0,
        "blocked_other_status": 0,
        "blocked_import_validation": 0,
        "blocked_duplicate_safeguard": 0,
        "created": 0,
        "dry_run_ready": 0,
        "api_errors": 0,
    }

    for row in rows:
        status_norm = normalize_status(row.get("Status", ""))

        if status_norm == STATUS_REVIEW_DUPLICATE:
            summary["blocked_review_duplicate"] += 1
            continue
        if status_norm == STATUS_REJECTED_INVALID:
            summary["blocked_rejected_invalid"] += 1
            continue
        if status_norm != STATUS_INSERT_READY:
            summary["blocked_other_status"] += 1
            continue

        summary["eligible_insert_ready"] += 1
        if args.limit and summary["eligible_insert_ready"] > args.limit:
            break

        payload, errors = build_payload(row, category_by_name)
        if errors:
            summary["blocked_import_validation"] += 1
            if args.verbose:
                print(f"[skip] row={row['_row']} validation={','.join(errors)}")
            continue

        dup = duplicate_signal(payload, existing_suppliers)
        if dup:
            summary["blocked_duplicate_safeguard"] += 1
            if args.verbose:
                print(f"[skip] row={row['_row']} duplicate={dup}")
            continue

        if not args.apply:
            summary["dry_run_ready"] += 1
            if args.verbose:
                print(f"[dry-run] row={row['_row']} payload-ready")
            continue

        code, res = req(args.base_url, args.auth_token, "POST", "/api/suppliers", payload)
        if code in (200, 201) and res.get("ok"):
            summary["created"] += 1
            existing_suppliers.append(payload)
            if args.verbose:
                print(f"[created] row={row['_row']} id={res.get('id')}")
        else:
            summary["api_errors"] += 1
            print(f"[error] row={row['_row']} status={code} body={res}")

    print(json.dumps({
        "mode": "apply" if args.apply else "dry-run",
        "workbook": workbook_path,
        "base_url": args.base_url,
        "summary": summary,
    }, indent=2))


if __name__ == "__main__":
    main()
